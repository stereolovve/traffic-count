import logging
import os
import pandas as pd
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.config import get_excel_dir

class ExcelManager:
    def __init__(self, contador):
        self.contador = contador
        self.session = contador.session

    def save_contagens(self, current_timeslot):
        try:
            # Verificar diretórios primeiro
            self._ensure_directories()
            arquivo_sessao = self._get_excel_path()
            
            # Verificar se o arquivo existe e criar se necessário
            if not os.path.exists(arquivo_sessao):
                logging.info(f"Arquivo Excel não encontrado. Criando novo: {arquivo_sessao}")
                if not self.initialize_excel_file():
                    raise Exception("Falha ao criar arquivo Excel")
            
            # Calcular período fim
            periodo_fim = current_timeslot + timedelta(minutes=15)
            periodo_str = current_timeslot.strftime("%H:%M")
            periodo_fim_str = periodo_fim.strftime("%H:%M")
            
            # Salvar no Excel
            with pd.ExcelWriter(arquivo_sessao, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for movimento in self.contador.details["Movimentos"]:
                    try:
                        df_existente = pd.read_excel(arquivo_sessao, sheet_name=movimento)
                    except Exception as ex:
                        logging.error(f"Erro ao ler planilha {movimento}: {ex}")
                        continue

                    # Verificar se o período já existe
                    if "das" in df_existente.columns and any(df_existente["das"] == periodo_str):
                        logging.warning(f"Período {periodo_str} já existe no Excel. Pulando...")
                        continue

                    nova_linha = {
                        "das": periodo_str,
                        "às": periodo_fim_str,
                        # Observação específica por movimento
                        "observacao": self.contador.period_observacoes.get(movimento, "")
                    }

                    # Adicionar contagens para cada veículo
                    for categoria in [c for c in self.contador.categorias if c.movimento == movimento]:
                        veiculo = categoria.veiculo
                        key = (veiculo, movimento)
                        nova_linha[veiculo] = self.contador.contagens.get(key, 0)

                    df_novo = pd.DataFrame([nova_linha])
                    df_resultante = pd.concat([df_existente, df_novo], ignore_index=True)
                    df_resultante.to_excel(writer, sheet_name=movimento, index=False)

            logging.info(f"Contagens salvas no Excel para o período: {periodo_str}")
            return True

        except Exception as ex:
            logging.error(f"Erro ao salvar contagens no Excel: {ex}")
            return False

    def initialize_excel_file(self):
        """Inicializa o arquivo Excel para uma nova sessão"""
        try:
            # 1. Verificar se temos os dados necessários
            if not self.contador.sessao:
                raise ValueError("Sessão não inicializada")
            
            if not self.contador.details.get("Movimentos"):
                raise ValueError("Movimentos não definidos")

            # 2. Criar diretórios
            self._ensure_directories()
            arquivo_sessao = self._get_excel_path()

            logging.info(f"Criando arquivo Excel: {arquivo_sessao}")

            # 3. Remover arquivo existente se houver
            if os.path.exists(arquivo_sessao):
                os.remove(arquivo_sessao)
                logging.info("Arquivo existente removido")

            # 4. Criar novo arquivo
            wb = Workbook()

            # 5. Configurar planilhas
            if not self.contador.details["Movimentos"]:
                ws = wb.active
                ws.title = "Placeholder"
                ws.append(["Aviso", "Nenhum movimento foi definido."])
                logging.warning("Nenhum movimento definido. Placeholder criado.")
            else:
                wb.remove(wb.active)
                for movimento in self.contador.details["Movimentos"]:
                    ws = wb.create_sheet(title=movimento)
                    # Adicionar cabeçalhos padrão
                    headers = ["das", "às", "observacao"]
                    # Adicionar veículos como colunas na ordem original (por ID)
                    veiculos_movimento = []
                    for categoria in self.contador.categorias:
                        if categoria.movimento == movimento and categoria.veiculo not in veiculos_movimento:
                            veiculos_movimento.append(categoria.veiculo)
                    headers.extend(veiculos_movimento)
                    ws.append(headers)
                    logging.info(f"Planilha criada para movimento: {movimento} com colunas: {headers}")

            # 6. Adicionar planilha de detalhes
            ws_details = wb.create_sheet(title="Detalhes")
            details_df = pd.DataFrame([self.contador.details])

            for coluna in details_df.columns:
                details_df[coluna] = details_df[coluna].apply(
                    lambda x: ', '.join(x) if isinstance(x, list) else x
                )

            for row in dataframe_to_rows(details_df, index=False, header=True):
                ws_details.append(row)

            # 7. Salvar arquivo
            wb.active = 0
            wb.save(arquivo_sessao)
            logging.info(f"Arquivo Excel inicializado com sucesso: {arquivo_sessao}")
            return True

        except Exception as ex:
            logging.error(f"Erro ao inicializar arquivo Excel: {ex}")
            return False

    def _get_excel_path(self):
        """Retorna o caminho do arquivo Excel"""
        nome_pesquisador = re.sub(r'[<>:"/\\|?*]', '', self.contador.username)
        codigo = re.sub(r'[<>:"/\\|?*]', '', self.contador.details['Código'])
        excel_dir = get_excel_dir()
        diretorio_pesquisador_codigo = os.path.join(excel_dir, nome_pesquisador, codigo)
        return os.path.join(diretorio_pesquisador_codigo, f'{self.contador.sessao}.xlsx')

    def _ensure_directories(self):
        """Garante que os diretórios necessários existam"""
        try:
            nome_pesquisador = re.sub(r'[<>:"/\\|?*]', '', self.contador.username)
            codigo = re.sub(r'[<>:"/\\|?*]', '', self.contador.details['Código'])
            
            excel_dir = get_excel_dir()
            
            diretorio_pesquisador = os.path.join(excel_dir, nome_pesquisador)
            diretorio_pesquisador_codigo = os.path.join(diretorio_pesquisador, codigo)
            
            # Criar diretórios com mensagens de log
            if not os.path.exists(diretorio_pesquisador):
                logging.info(f"Criando diretório do pesquisador: {diretorio_pesquisador}")
                os.makedirs(diretorio_pesquisador)
            
            if not os.path.exists(diretorio_pesquisador_codigo):
                logging.info(f"Criando diretório do código: {diretorio_pesquisador_codigo}")
                os.makedirs(diretorio_pesquisador_codigo)
            
            return True
            
        except Exception as ex:
            logging.error(f"Erro ao criar diretórios: {ex}")
            raise
