import flet as ft
from database.models import Session, Categoria, Sessao, Contagem, Historico, init_db
from sqlalchemy.exc import SQLAlchemyError
from app import aba_inicio, aba_contagem, aba_historico, listener, sessao
from utils.initializer import inicializar_variaveis, configurar_numpad_mappings
from auth.login import LoginPage
from utils.period import format_period
from auth.register import RegisterPage
from pynput import keyboard
from utils.padrao_contagem import carregar_categorias_padrao, carregar_padroes_selecionados, obter_caminho_json
import pandas as pd
from datetime import datetime, timedelta, time
import os
import re
from utils.change_binds import change_binds
import logging
import httpx
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json

#------------------------ LOGGING -------------------------
logging.basicConfig(
    level=logging.ERROR, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('log.txt'), 
              logging.StreamHandler()]
)
#------------------------ SETUP DATABASE ------------------
init_db()
#------------------------ CONTADOR PERPLAN ----------------
class ContadorPerplan(ft.Column):
    #------------------------ CONSTRUTOR ------------------
    def __init__(self, page, username, app):
        
        super().__init__()
        self.tokens = None
        self.page = page
        self.username = username
        self.app = app
        inicializar_variaveis(self)  
        configurar_numpad_mappings(self) 
        self.setup_ui()
        self.page.update()
        self.carregar_sessao_ativa()
        self.pressed_keys = set()
    #------------------------ SETUP UI ------------------------
    def setup_ui(self):
        self.tabs = ft.Tabs(
            tabs=[
                ft.Tab(text="Inicio", content=ft.Column(width=450, height=900)),
                ft.Tab(text="Contador", content=ft.Column(width=450, height=900)),
                ft.Tab(text="Histórico", content=ft.Column(width=450, height=900)),
                ft.Tab(text="", icon=ft.icons.SETTINGS, content=ft.Column(width=450, height=900))
            ]
        )
        self.controls.clear()
        self.controls.append(self.tabs)
        self.setup_aba_inicio()
        self.setup_aba_contagem()
        self.setup_aba_historico()
        self.setup_aba_config()
        self.tabs.tabs[1].content.visible = False
        self.atualizar_borda_contagem()
    #------------------------ ABA INICIO ------------------------
    def setup_aba_inicio(self):
        tab = self.tabs.tabs[0].content
        tab.controls.clear()
        hoje_str = datetime.now().strftime("%d-%m-%Y")
        
        ghost_label = ft.Text("Código", color=ft.colors.TRANSPARENT, size=15, weight=ft.FontWeight.W_400)
        self.codigo_ponto_input = ft.TextField(label="Código", hint_text="exemplo: ER2403", icon=ft.icons.CODE)
        self.nome_ponto_input = ft.TextField(label="Ponto", hint_text="exemplo: P10; P15", icon=ft.icons.LOCATION_PIN)
        self.data_ponto_input = ft.TextField(label="Data", hint_text="dd-mm-aaaa", value=hoje_str, icon=ft.icons.CALENDAR_MONTH, keyboard_type=ft.KeyboardType.NUMBER)
        self.selected_time = "00:00"

        def picker_changed(e):
            selected_seconds = e.control.value
            hours, remainder = divmod(selected_seconds, 3600)
            minutes = (remainder // 60)
            adjusted_minutes = (minutes // 15) * 15
            self.selected_time = f"{hours:02}:{adjusted_minutes:02}"
            self.time_picker_button.text = f"{self.selected_time}"
            self.time_picker_button.update()


        self.time_picker = ft.CupertinoTimerPicker(
            mode=ft.CupertinoTimerPickerMode.HOUR_MINUTE,
            value=0,
            minute_interval=15,
            on_change=picker_changed
        )

        self.time_picker_button = ft.ElevatedButton(
            text=f"{self.selected_time}",
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=1)),
            width=float('inf'),
            height=40,
            elevation=0,
            icon=ft.icons.ACCESS_ALARM,
            on_click=lambda _: self.page.open(
                ft.AlertDialog(content=self.time_picker)
            )
        )

        self.padrao_dropdown = ft.Dropdown(
            label="Selecione o padrão",
            options=[
                ft.dropdown.Option("Padrão Perplan"),
                ft.dropdown.Option("Padrão Perci"),
                ft.dropdown.Option("Padrão Simplificado"),
            ],
            on_change=self.carregar_padroes_selecionados,
        )

        self.movimentos_container = ft.Column()

        def adicionar_campo_movimento(e):
            movimento_input = ft.TextField(label="Movimento", hint_text="Exemplo: A, B")
            remover_button = ft.IconButton(
                icon=ft.icons.REMOVE,
                on_click=lambda _: remover_campo(movimento_input, remover_button)
            )
            movimento_row = ft.Row(controls=[movimento_input, remover_button], spacing=10)
            self.movimentos_container.controls.append(movimento_row)
            self.page.update()

        def remover_campo(movimento_input, remover_button):
            movimento_row = next(
                (row for row in self.movimentos_container.controls if movimento_input in row.controls),
                None
            )
            if movimento_row:
                self.movimentos_container.controls.remove(movimento_row)
                self.page.update()

        adicionar_movimento_button = ft.ElevatedButton(
            text="Adicionar Movimento",
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=1)),
            icon=ft.icons.ADD,
            width=float('inf'),
            height=40,
            on_click=adicionar_campo_movimento
        )

        criar_sessao_button = ft.ElevatedButton(
            text="Criar Sessão", on_click=self.criar_sessao,
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=1)),
            width=float('inf'),
            height=40,
            icon=ft.icons.SEND
        )

        self.sessao_status = ft.Text("Nenhuma sessão ativa", weight=ft.FontWeight.BOLD, size=20)

        tab.controls.extend([
            ghost_label,
            self.codigo_ponto_input,
            self.nome_ponto_input,
            self.data_ponto_input,
            self.time_picker_button,
            self.padrao_dropdown,
            adicionar_movimento_button,
            self.movimentos_container,
            criar_sessao_button,
            self.sessao_status
        ])

        self.page.update()


    def validar_campos(self):
        campos_obrigatorios = [
            (self.codigo_ponto_input, "Código"),
            (self.nome_ponto_input, "Ponto"),
            (self.time_picker_button, "Horário de Início"),
            (self.data_ponto_input, "Data do Ponto")
        ]

        for campo, nome in campos_obrigatorios:
            if isinstance(campo, ft.TextField) and not campo.value:
                snackbar = ft.SnackBar(ft.Text(f"{nome} é obrigatório!"), bgcolor="ORANGE")
                self.page.overlay.append(snackbar)
                snackbar.open = True
                self.page.update()
                return False
            elif isinstance(campo, ft.ElevatedButton) and campo.text == "Selecionar Horário":
                snackbar = ft.SnackBar(ft.Text(f"{nome} é obrigatório!"), bgcolor="ORANGE")
                self.page.overlay.append(snackbar)
                snackbar.open = True
                self.page.update()
                return False

        return True


    def criar_sessao(self, e):
        sessao.criar_sessao(self, e)


    def _inicializar_arquivo_excel(self):
        try:
            nome_pesquisador = re.sub(r'[<>:"/\\|?*]', '', self.username)
            codigo = re.sub(r'[<>:"/\\|?*]', '', self.details['Código'])
            diretorio_base = r'Z:\\0Pesquisa\\_0ContadorDigital\\Contagens'

            if not os.path.exists(diretorio_base):
                os.makedirs(diretorio_base, exist_ok=True)

            diretorio_pesquisador_codigo = os.path.join(diretorio_base, nome_pesquisador, codigo)
            if not os.path.exists(diretorio_pesquisador_codigo):
                os.makedirs(diretorio_pesquisador_codigo, exist_ok=True)

            arquivo_sessao = os.path.join(diretorio_pesquisador_codigo, f'{self.sessao}.xlsx')

            if os.path.exists(arquivo_sessao):
                os.remove(arquivo_sessao)

            wb = Workbook()

            if not self.details["Movimentos"]:
                ws = wb.active
                ws.title = "Placeholder"
                ws.append(["Aviso", "Nenhum movimento foi definido."])
                logging.warning("Nenhum movimento definido. Placeholder criado.")
            else:
                wb.remove(wb.active)
                for movimento in self.details["Movimentos"]:
                    wb.create_sheet(title=movimento)

            ws_details = wb.create_sheet(title="Detalhes")
            details_df = pd.DataFrame([self.details])

            for coluna in details_df.columns:
                details_df[coluna] = details_df[coluna].apply(
                    lambda x: ', '.join(x) if isinstance(x, list) else x
                )

            for row in dataframe_to_rows(details_df, index=False, header=True):
                ws_details.append(row)

            wb.active = 0

            wb.save(arquivo_sessao)
            logging.info(f"Arquivo Excel inicializado com sucesso: {arquivo_sessao}")

        except Exception as ex:
            logging.error(f"Erro ao inicializar arquivo Excel: {ex}")
            raise


    def confirmar_finalizar_sessao(self, e):
        sessao.confirmar_finalizar_sessao(self, e)

    def end_session(self):
        sessao.end_session(self)

    # ------------------------ ABA CONTADOR ------------------------
    def setup_aba_contagem(self):
        aba_contagem.setup_aba_contagem(self)
        
    def resetar_todas_contagens(self, e):
        aba_contagem.resetar_todas_contagens(self, e)

    def confirmar_resetar_todas_contagens(self, e):
        aba_contagem.confirmar_resetar_todas_contagens(self, e)

    def create_moviment_content(self, movimento):
        content = ft.Column()
        categorias = [c for c in self.categorias if c.movimento == movimento]
        for categoria in categorias:
            control = self.create_category_control(categoria.veiculo, categoria.bind, movimento)
            content.controls.append(control)

        return content


    def create_category_control(self, veiculo, bind, movimento):
        label_veiculo = ft.Text(f"{veiculo}", size=15, width=100)
        label_bind = ft.Text(f"({bind})", color="cyan", size=15, width=50)
        label_count = ft.Text(f"{self.contagens.get((veiculo, movimento), 0)}", size=15, width=50)
        self.labels[(veiculo, movimento)] = label_count

        campo_visivel = False

        popup_menu = ft.PopupMenuButton(
            icon_color="teal",
            items=[
                ft.PopupMenuItem(text="Adicionar", icon=ft.icons.ADD, on_click=lambda e, v=veiculo, m=movimento: self.increment(v, m)),
                ft.PopupMenuItem(text="Remover", icon=ft.icons.REMOVE, on_click=lambda e, v=veiculo, m=movimento: self.decrement(v, m)),
                ft.PopupMenuItem(text="Editar Contagem", icon=ft.icons.EDIT, on_click=lambda e: self.abrir_edicao_contagem(veiculo, movimento)),
            ]
        )

        return ft.Row(
            alignment=ft.MainAxisAlignment.CENTER,
            spacing=5,
            controls=[
                ft.Container(content=label_veiculo, alignment=ft.alignment.center_left),
                ft.Container(content=label_bind, alignment=ft.alignment.center),
                ft.Container(content=label_count, alignment=ft.alignment.center_right),
                popup_menu
            ]
        )


    def toggle_contagem(self, e):
        self.contagem_ativa = e.control.value
        self.atualizar_borda_contagem()
        self.page.update()

    def atualizar_borda_contagem(self):
        aba_contagem.atualizar_borda_contagem(self)

    def increment(self, veiculo, movimento):
        try:
            self.contagens[(veiculo, movimento)] = self.contagens.get((veiculo, movimento), 0) + 1
            self.update_labels(veiculo, movimento)
            self.save_to_db(veiculo, movimento)
            self.salvar_historico(veiculo, movimento, "incremento")
            self.update_current_tab()
            self.page.update()

        except Exception as ex:
            logging.error(f"Erro ao incrementar: {ex}")

    def decrement(self, veiculo, movimento):
        try:
            if self.contagens.get((veiculo, movimento), 0) > 0:
                self.contagens[(veiculo, movimento)] = self.contagens.get((veiculo, movimento), 0) - 1
                self.update_labels(veiculo, movimento)
                self.save_to_db(veiculo, movimento)
                self.salvar_historico(veiculo, movimento, "remoção")
                self.page.update()

        except Exception as ex:
            logging.error(f"Erro ao decrementar: {ex}")


    def abrir_edicao_contagem(self, veiculo, movimento):
        self.contagem_ativa = False
        self.atualizar_borda_contagem()
        self.page.update()
        def on_submit(e):
            try:
                nova_contagem = int(input_contagem.value)
                self.contagens[(veiculo, movimento)] = nova_contagem
                self.update_labels(veiculo, movimento)
                self.save_to_db(veiculo, movimento)
                self.salvar_historico(veiculo, movimento, "edição manual")
                snackbar = ft.SnackBar(ft.Text(f"Contagem de '{veiculo}' no movimento '{movimento}' foi atualizada para {nova_contagem}."), bgcolor="BLUE")
                self.page.overlay.append(snackbar)
                snackbar.open = True
                dialog.open = False
                self.contagem_ativa = True
                self.page.update()

            except ValueError:
                logging.error("[ERROR] Valor de contagem inválido.")

        input_contagem = ft.TextField(label="Nova Contagem", keyboard_type=ft.KeyboardType.NUMBER, on_submit=on_submit)
        dialog = ft.AlertDialog(
            title=ft.Text(f"Editar Contagem: {veiculo}"),
            content=input_contagem,
            actions=[ft.TextButton("Salvar", on_click=on_submit)],
            on_dismiss=lambda e: dialog.open == False,
        )
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()


    def salvar_historico(self, veiculo, movimento, acao):
        try:
            novo_registro = Historico(
                sessao=self.sessao,
                veiculo=veiculo,
                movimento=movimento,
                timestamp=datetime.now(),
                acao=acao
            )
            self.session.add(novo_registro)
            self.session.commit()
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao salvar histórico: {ex}")
            self.session.rollback()


    def update_current_tab(self):
        current_tab = self.movimento_tabs.tabs[self.movimento_tabs.selected_index]
        current_tab.content.update()
        self.page.update()

    def update_labels(self, veiculo, movimento):
            self.labels[(veiculo, movimento)].value = str(self.contagens.get((veiculo, movimento), 0))
            self.page.update()
            
    def update_sessao_status(self):
        self.sessao_status.value = f"Sessão ativa: {self.sessao}" if self.sessao else "Nenhuma sessão ativa"
        self.page.update()


    def save_contagens(self, e):
        try:
            now = datetime.now()
            if hasattr(self, 'last_save_time'):
                time_since_last_save = now - self.last_save_time
                if time_since_last_save < timedelta(minutes=5):
                    def on_confirm_save(e):
                        dialog.open = False
                        self.page.update()
                        self._perform_save(now)
                    def on_cancel_save(e):
                        dialog.open = False
                        self.page.update()

                    dialog = ft.AlertDialog(
                        title=ft.Text("Confirmar Salvamento"),
                        content=ft.Text("Você salvou recentemente. Deseja salvar novamente?"),
                        actions=[
                            ft.TextButton("Sim", on_click=on_confirm_save),
                            ft.TextButton("Cancelar", on_click=on_cancel_save),
                        ],
                    )
                    self.page.overlay.append(dialog)
                    dialog.open = True
                    self.page.update()
                    return
            self._perform_save(now)
            self.update_last_save_label(now)

            snackbar = ft.SnackBar(ft.Text("Contagens salvas com sucesso!"), bgcolor="GREEN")
            self.page.overlay.append(snackbar)
            snackbar.open = True
        except Exception as ex:
            logging.error(f"Erro ao salvar contagens: {ex}")
            snackbar = ft.SnackBar(ft.Text("Erro ao salvar contagens."), bgcolor="RED")
            self.page.overlay.append(snackbar)
            snackbar.open = True
            self.page.update()

    def _perform_save(self, now):
        try:
            horario_atual = self.current_timeslot

            nome_pesquisador = re.sub(r'[<>:"/\\|?*]', '', self.username)
            codigo = re.sub(r'[<>:"/\\|?*]', '', self.details['Código'])
            diretorio_base = r'Z:\\0Pesquisa\\_0ContadorDigital\\Contagens'
            diretorio_pesquisador_codigo = os.path.join(diretorio_base, nome_pesquisador, codigo)
            arquivo_sessao = os.path.join(diretorio_pesquisador_codigo, f'{self.sessao}.xlsx')

            with pd.ExcelWriter(arquivo_sessao, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for movimento in self.details["Movimentos"]:
                    df_existente = pd.read_excel(arquivo_sessao, sheet_name=movimento)

                    nova_linha = {
                        "das": horario_atual.strftime("%H:%M"),
                        "às": (horario_atual + timedelta(minutes=15)).strftime("%H:%M"),
                    }

                    for categoria in [c for c in self.categorias if c.movimento == movimento]:
                        nova_linha[categoria.veiculo] = self.contagens.get((categoria.veiculo, movimento), 0)

                    df_novo = pd.DataFrame([nova_linha])
                    df_resultante = pd.concat([df_existente, df_novo], ignore_index=True)

                    df_resultante.to_excel(writer, sheet_name=movimento, index=False)

            self.current_timeslot += timedelta(minutes=15)
            self.details["current_timeslot"] = self.current_timeslot.strftime("%H:%M")
            self.salvar_sessao()
            self.last_save_time = now
            self.update_last_save_label(now)

            logging.info(f"Salvamento realizado com sucesso: {arquivo_sessao}")

        except Exception as ex:
            logging.error(f"Erro ao salvar contagens: {ex}")
            raise
    def update_last_save_label(self, now):
        if hasattr(self, "last_save_label"):
            self.last_save_label.value = f"Último salvamento: {now.strftime('%H:%M:%S')}"
            self.last_save_label.update()
        else:
            self.last_save_label = ft.Text(
                f"Último salvamento: {now.strftime('%H:%M:%S')}",
                size=16,
                weight=ft.FontWeight.W_400,
                alignment=ft.alignment.center_left
            )
            self.tabs.tabs[1].content.controls.insert(0, self.last_save_label)
            self.page.update()



    def confirmar_finalizar_sessao(self, e):
        def close_dialog(e):
            dialog.open = False
            self.page.update()

        def end_and_close(e):
            try:
                dialog.open = False
                self.page.update()
                self.end_session()
            except Exception as ex:
                logging.error(f"Erro ao finalizar sessão: {ex}")
            
        dialog = ft.AlertDialog(
            title=ft.Text("Finalizar Sessão"),
            content=ft.Text("Você tem certeza que deseja finalizar a sessão?"),
            actions=[
                ft.TextButton("Sim", on_click=end_and_close),
                ft.TextButton("Cancelar", on_click=close_dialog),
            ],
        )
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()

    def end_session(self):
        try:
            self.finalizar_sessao()
            
            for veiculo, movimento in list(self.contagens.keys()):
                self.contagens[(veiculo, movimento)] = 0
                self.update_labels(veiculo, movimento)
                self.save_to_db(veiculo, movimento)
            
            self.sessao = None
            self.details = {"Movimentos": []}
            self.contagens = {}
            self.binds = {}
            self.labels = {}

            self.page.overlay.append(ft.SnackBar(ft.Text("Sessão finalizada!")))
            self.page.update()

            self.restart_app()
        except Exception as ex:
            logging.error(f"Erro ao finalizar sessão: {ex}")
            self.page.overlay.append(ft.SnackBar(ft.Text(f"Erro ao finalizar sessão: {ex}")))
            self.page.update()


    def restart_app(self):
        self.stop_listener()
        self.sessao = None
        self.details = {"Movimentos": []}
        self.contagens = {}
        self.binds = {}
        self.labels = {}
        self.setup_ui()
        self.page.update()
        self.start_listener()

    def save_to_db(self, veiculo, movimento):
        try:
            contagem = self.session.query(Contagem).filter_by(sessao=self.sessao, veiculo=veiculo, movimento=movimento).first()
            if contagem:
                contagem.count = self.contagens.get((veiculo, movimento), 0)
            else:
                nova_contagem = Contagem(
                    sessao=self.sessao,
                    veiculo=veiculo,
                    movimento=movimento,
                    count=self.contagens.get((veiculo, movimento), 0)
                )
                self.session.add(nova_contagem)
            self.session.commit()
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao salvar no DB: {ex}")
            self.session.rollback()

    def recuperar_contagens(self):
        try:
            contagens_db = self.session.query(Contagem).filter_by(sessao=self.sessao).all()
            for contagem in contagens_db:
                self.contagens[(contagem.veiculo, contagem.movimento)] = contagem.count
                self.update_labels(contagem.veiculo, contagem.movimento)
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao recuperar contagens: {ex}")

    def atualizar_atalho(self, e, veiculo, movimento):
        novo_atalho = e.control.value
        try:
            categoria = self.session.query(Categoria).filter_by(veiculo=veiculo, movimento=movimento).first()
            if categoria:
                categoria.bind = novo_atalho
                self.session.commit()
                self.update_binds()
                snackbar = ft.SnackBar(ft.Text(f"Atalho atualizado para {veiculo}"), bgcolor="GREEN")
                self.page.overlay.append(snackbar)
                snackbar.open = True
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao atualizar atalho: {ex}")
            self.session.rollback()
        self.page.update()

    def atualizar_bind(self, e, veiculo, movimento):
        novo_bind = e.control.value
        try:
            categoria = self.session.query(Categoria).filter_by(veiculo=veiculo, movimento=movimento).first()
            if categoria:
                categoria.bind = novo_bind
                self.session.commit()
                self.update_binds()
                snackbar = ft.SnackBar(ft.Text(f"Bind atualizado para {veiculo}"), bgcolor="GREEN")
                self.page.overlay.append(snackbar)
                snackbar.open = True
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao atualizar bind: {ex}")
            self.session.rollback()
        self.page.update()

    def editar_bind(self, veiculo, movimento):
        def on_bind_submit(e):
            new_bind = bind_input.value
            self.update_bind(veiculo, new_bind, movimento)
            dialog.open = False
            self.page.update()

        bind_input = ft.TextField(label="Novo Bind", width=100)
        dialog = ft.AlertDialog(
            title=ft.Text("Editar Bind"),
            content=bind_input,
            actions=[
                ft.TextButton("Salvar", on_click=on_bind_submit),
                ft.TextButton("Cancelar", on_click=lambda e: self.close_dialog(dialog)),
            ]
        )
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()

    def close_dialog(self, dialog):
        dialog.open = False
        self.page.update()

    def update_bind(self, veiculo, new_bind, movimento):
        if new_bind:
            try:
                categorias = self.session.query(Categoria).filter_by(veiculo=veiculo).all()
                for categoria in categorias:
                    categoria.bind = new_bind
                self.session.commit()
                self.update_binds()
                self.update_ui()
            except SQLAlchemyError as ex:
                logging.error(f"Erro ao atualizar bind: {ex}")
                self.session.rollback()

    def update_binds(self):
        try:
            self.binds = {(categoria.bind, categoria.movimento): (categoria.veiculo, categoria.movimento)
                        for categoria in self.session.query(Categoria).all()}
            logging.info("Binds atualizados com sucesso.")
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao atualizar binds: {ex}")

    def update_ui(self):
        self.setup_aba_contagem()
        self.page.update()

    #------------------- ABA HISTÓRICO -----------------------------
    def setup_aba_historico(self):
        tab = self.tabs.tabs[2].content
        tab.controls.clear()

        header = ft.Row(
            alignment=ft.MainAxisAlignment.CENTER,
            controls=[
                ft.Container(content=ft.Text("Histórico de Contagens", weight=ft.FontWeight.W_400, size=15))
            ]
        )

        self.historico_lista = ft.ListView(spacing=10, padding=20, auto_scroll=True)

        carregar_historico_button = ft.ElevatedButton(
            text="Carregar próximos 30 registros",
            on_click=self.carregar_historico
        )

        tab.controls.extend([
            header,
            carregar_historico_button,
            self.historico_lista
        ])
        self.page.update()

    def carregar_historico(self, e):
        aba_historico.carregar_historico(self, e)

    #------------------- ABA CONFIGURACOES -----------------------------
    def setup_aba_config(self):
        tab = self.tabs.tabs[3].content
        tab.controls.clear()

        avatar = ft.CircleAvatar(
           # foreground_image_url="https://example.com/imagem_perfil.png",  # Troque pela URL da imagem ou remova se não tiver uma
            radius=40,
        )
        
        username_text = ft.Text(
            f"Conectado como: {self.username}",
            weight=ft.FontWeight.W_400,
            size=15,
            text_align=ft.TextAlign.CENTER
        )

        profile_container = ft.Column(
            controls=[
                ft.Container(avatar, alignment=ft.alignment.center),
                ft.Container(username_text, alignment=ft.alignment.center),
            ],
            spacing=5,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER
        )

        self.page.theme_mode = ft.ThemeMode.DARK
        self.modo_claro_escuro = ft.Switch(label="Modo claro", on_change=self.theme_changed)

        opacity = ft.Slider(value=100, min=20, max=100, divisions=20, label="Opacidade", on_change=self.ajustar_opacidade)
        
        config_button = ft.ElevatedButton(
            text="Configurar Binds", 
            on_click=lambda e: change_binds(self.page, self),
            icon=ft.icons.SETTINGS
        )
        
        logout_button = ft.ElevatedButton(
            text="Sair",
            bgcolor="RED",
            color="WHITE",
            on_click=self.logout_user,
            icon=ft.icons.LOGOUT
        )

        config_layout = ft.Column(
            controls=[
                profile_container,
                ft.Divider(),
                ft.Text("Aparência", weight=ft.FontWeight.BOLD, size=16),
                self.modo_claro_escuro,
                ft.Divider(),
                ft.Text("Transparência da Janela", weight=ft.FontWeight.BOLD, size=16),
                opacity,
                ft.Divider(),
                ft.Text("Configurações Avançadas", weight=ft.FontWeight.BOLD, size=16),
                config_button,
                ft.Divider(),
                ft.Text("Deslogar:", weight=ft.FontWeight.BOLD, size=16),
                logout_button
            ],
            spacing=20,
            horizontal_alignment=ft.CrossAxisAlignment.STRETCH,
            scroll=ft.ScrollMode.AUTO
        )

        tab.controls.append(config_layout)
        self.page.update()

    def theme_changed(self, e):
        self.page.theme_mode = (
            ft.ThemeMode.DARK
            if self.page.theme_mode == ft.ThemeMode.LIGHT
            else ft.ThemeMode.LIGHT
        )
        self.modo_claro_escuro.label = (
            "Modo claro" if self.page.theme_mode == ft.ThemeMode.LIGHT else "Modo escuro"
        )
        self.page.update()

    def ajustar_opacidade(self, e):
        try:
            nova_opacidade = e.control.value / 100
            self.page.window.opacity = nova_opacidade
            self.page.update()
        except Exception as ex:
            logging.error(f"Erro ao ajustar opacidade: {ex}")

    #------------------- LISTENER -----------------------------
    def on_key_press(self, key):
        listener.on_key_press(self, key)

    def on_key_release(self, key):
        try:
            self.pressed_keys.discard(key)
        except Exception as ex:
            logging.error(f"Erro no on_key_release: {ex}")

    def start_listener(self):
        if self.listener is None:
            self.listener = keyboard.Listener(on_press=self.on_key_press, on_release=self.on_key_release)
            self.listener.start()

    def stop_listener(self):
        if self.listener is not None:
            self.listener.stop()
            self.listener = None
            self.pressed_keys.clear()

    def carregar_sessao_ativa(self):
        sessao.carregar_sessao_ativa(self)

    def salvar_sessao(self):
        sessao.salvar_sessao(self)

    def finalizar_sessao(self):
        sessao.finalizar_sessao(self)

    def carregar_categorias_padrao(self, caminho_json, padrao):
        try:
            with open(caminho_json, 'r') as f:
                categorias_padrao = json.load(f)

            if not categorias_padrao:
                raise Exception(f"Nenhuma categoria encontrada no arquivo JSON: {caminho_json}")

            for categoria in categorias_padrao:
                veiculo = categoria.get('veiculo')
                bind = categoria.get('bind')

                if veiculo and bind:
                    for movimento in self.details.get("Movimentos", []):
                        categoria_existente = self.session.query(Categoria).filter_by(
                            veiculo=veiculo,
                            movimento=movimento
                        ).first()

                        if categoria_existente:
                            categoria_existente.bind = bind
                            categoria_existente.padrao = padrao
                            logging.info(f"Categoria atualizada: {veiculo} - {movimento} ({bind})")
                        else:
                            nova_categoria = Categoria(
                                padrao=padrao,
                                veiculo=veiculo,
                                movimento=movimento,
                                bind=bind,
                                criado_em=datetime.now()
                            )
                            self.session.add(nova_categoria)
                            logging.info(f"Categoria adicionada: {veiculo} - {movimento} ({bind})")

            self.session.commit()
            logging.info(f"Categorias do padrão '{padrao}' carregadas com sucesso.")

        except FileNotFoundError:
            logging.error(f"Arquivo JSON não encontrado: {caminho_json}")
            raise Exception("Arquivo JSON não encontrado.")
        except json.JSONDecodeError:
            logging.error(f"Erro ao decodificar JSON: {caminho_json}")
            raise Exception("Erro ao decodificar JSON.")
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao salvar padrões no banco de dados: {ex}")
            self.session.rollback()
            raise Exception("Erro ao salvar no banco de dados.")
        except Exception as ex:
            logging.error(f"Erro inesperado ao carregar categorias: {ex}")
            raise Exception(f"Erro ao carregar categorias: {ex}")

    def carregar_padroes_selecionados(self, e=None):
        try:
            padrao_selecionado = self.padrao_dropdown.value
            if not padrao_selecionado:
                snackbar = ft.SnackBar(ft.Text("Selecione um padrão!"), bgcolor="ORANGE")
                self.page.overlay.append(snackbar)
                snackbar.open = True
                self.page.update()
                return

            self.session.query(Categoria).delete()
            self.session.commit()

            caminho_json = self.obter_caminho_json(padrao_selecionado)
            if not caminho_json:
                snackbar = ft.SnackBar(ft.Text(f"Padrão '{padrao_selecionado}' não encontrado!"), bgcolor="RED")
                self.page.overlay.append(snackbar)
                snackbar.open = True
                return

            self.carregar_categorias_padrao(caminho_json, padrao=padrao_selecionado)

            snackbar = ft.SnackBar(ft.Text(f"Padrão '{padrao_selecionado}' carregado com sucesso!"), bgcolor="GREEN")
            self.page.overlay.append(snackbar)
            snackbar.open = True
            self.page.update()
        except SQLAlchemyError as ex:
            self.session.rollback()
            logging.error(f"Erro ao carregar padrões no banco de dados: {ex}")
            snackbar = ft.SnackBar(ft.Text(f"Erro ao carregar padrões: {ex}"), bgcolor="RED")
            self.page.overlay.append(snackbar)
            snackbar.open = True
        except Exception as ex:
            logging.error(f"Erro ao carregar padrões: {ex}")
            snackbar = ft.SnackBar(ft.Text(f"Erro ao carregar padrões: {ex}"), bgcolor="RED")
            self.page.overlay.append(snackbar)
            snackbar.open = True

    def obter_caminho_json(self, padrao_selecionado):
        base_path = os.path.join(os.getcwd(), "utils")
        if padrao_selecionado == "Padrão Perplan":
            return os.path.join(base_path, "padrao_perplan.json")
        elif padrao_selecionado == "Padrão Perci":
            return os.path.join(base_path, "padrao_perci.json")
        elif padrao_selecionado == "Padrão Simplificado":
            return os.path.join(base_path, "padrao_simplificado.json")
        return None

    def carregar_config(self):
        try:
            categorias = self.session.query(Categoria).order_by(Categoria.criado_em).all()
            binds = {}
            for categoria in categorias:
                binds[(categoria.bind, categoria.movimento)] = (categoria.veiculo, categoria.movimento)
            
            contagens = {}
            if self.sessao:
                contagens_db = self.session.query(Contagem).filter_by(sessao=self.sessao).all()
                for contagem in contagens_db:
                    contagens[(contagem.veiculo, contagem.movimento)] = contagem.count
            else:
                logging.warning("Nenhuma sessão ativa para carregar contagens.")
            
            return contagens, binds, categorias
        except SQLAlchemyError as ex:
            logging.error(f"Erro ao carregar config: {ex}")
            return {}, {}, []

    def update_sessao_status(self):
        self.sessao_status.value = f"Sessão ativa: {self.sessao}" if self.sessao else "Nenhuma sessão ativa"
        self.page.update()
        
        
    def logout_user(self, e):
        try:
            if os.path.exists("auth_tokens.json"):
                os.remove("auth_tokens.json")

            if not self.page:
                logging.error("Página está faltando ao tentar deslogar.")
                return

            self.app.reset_app()
        except AttributeError as ex:
            logging.error(f"Erro ao deslogar: {ex}")
        

#------------------- MAIN -----------------------------
def main(page: ft.Page):
    page.title = "Contador Perplan"
    page.window.width = 800
    page.window.height = 600
    page.window.always_on_top = True
    page.scroll = ft.ScrollMode.AUTO
    page.window.center()
    class MyApp:
        def __init__(self, page):
            self.page = page
            self.tokens = None
            self.username = None

            if not self.page:
                logging.error("Página não está configurada ao iniciar o app.")
                return
        
        def switch_to_main_app(self):
            if not self.page:
                logging.error("Página não está disponível ao alternar para o aplicativo principal.")
                return
            self.page.controls.clear()
            contador = ContadorPerplan(page, username=self.username, app=self)
            self.page.add(contador)

            self.page.window.width = 800
            self.page.window.height = 700
            self.page.window.always_on_top = True
            self.page.scroll = ft.ScrollMode.AUTO
            self.page.update()
            contador.start_listener()

            self.page.on_close = lambda e: contador.stop_listener()
            self.page.update()

        def show_login_page(self):
            self.page.controls.clear()
            self.page.add(LoginPage(self))
            self.page.update()

        def show_register_page(self):
            self.page.controls.clear()
            self.page.add(RegisterPage(self))
            self.page.update()

        def reset_app(self):
            self.tokens = None
            self.username = None
            if not self.page:
                logging.error("Página não está configurada ao tentar resetar o app.")
                return
            self.page.controls.clear()
            self.show_login_page()
    app = MyApp(page)
    
    if os.path.exists("auth_tokens.json"):
        try:
            with open("auth_tokens.json", "r") as f:
                saved_data = json.load(f)
                app.tokens = saved_data["tokens"]
                app.username = saved_data["username"]
                app.switch_to_main_app()
        except Exception as ex:
            logging.error(f"Erro ao carregar tokens salvos: {ex}")
            app.show_login_page()
    else:
        app.show_login_page()
        
ft.app(target=main)