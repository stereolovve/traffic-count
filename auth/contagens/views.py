#auth/contagens/views.py
from django.shortcuts import render, get_object_or_404
from .models import Session, Counting
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, FileResponse
import json
from collections import defaultdict
from django.contrib.auth import get_user_model
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
from padroes.models import PadraoContagem
from django.core.paginator import Paginator
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
from django.views.decorators.http import require_http_methods
import json
import logging

User = get_user_model()

def listar_sessoes(request):
    # Get filter parameters
    status = request.GET.get("status")
    pesquisador = request.GET.get("usuario")
    codigo = request.GET.get("codigo")
    ponto = request.GET.get("ponto")
    data = request.GET.get("data")
    padrao = request.GET.get("padrao")
    id_sessao = request.GET.get("id", '')
    sort_field = request.GET.get("sort", "-id")
    page = request.GET.get('page', 1)

    # Get list of researchers (users who have created sessions)
    pesquisadores = User.objects.filter(
        id__in=Session.objects.values_list('criado_por', flat=True).distinct()
    ).order_by('username')

    # Get unique codes and points
    codigos = Session.objects.values_list('codigo', flat=True).distinct().order_by('codigo')
    pontos = Session.objects.values_list('ponto', flat=True).distinct().order_by('ponto')
    datas = Session.objects.values_list('data', flat=True).distinct().order_by('data')
    
    padroes = PadraoContagem.objects.values_list('pattern_type', flat=True).distinct().order_by('pattern_type')

    sessoes = Session.objects.all().select_related('criado_por')

    if status == "aguardando":
        sessoes = sessoes.filter(status="Aguardando")
    elif status == "em_andamento":
        sessoes = sessoes.filter(status="Em andamento")
    elif status == "concluida":
        sessoes = sessoes.filter(status="Concluída")

    if pesquisador:
        sessoes = sessoes.filter(criado_por__username__icontains=pesquisador)

    if codigo:
        sessoes = sessoes.filter(codigo__icontains=codigo)

    if ponto:
        sessoes = sessoes.filter(ponto__icontains=ponto)

    if data:
        sessoes = sessoes.filter(data__icontains=data)
        
    if padrao:
        sessoes = sessoes.filter(padrao=padrao)
        
    if id_sessao:
        try:
            sessoes = sessoes.filter(id=int(id_sessao))
        except ValueError:
            pass

    if sort_field:
        if sort_field == 'criado_por':
            sessoes = sessoes.order_by(f'criado_por__username{"" if sort_field[0] != "-" else ""}')
        elif sort_field == 'created_at' or sort_field == '-created_at':
            sessoes = sessoes.order_by(sort_field)
        elif sort_field == 'updated_at' or sort_field == '-updated_at':
            sessoes = sessoes.order_by(sort_field)
        else:
            sessoes = sessoes.order_by(sort_field)

    sessoes_com_movimentos = []
    for sessao in sessoes:
        movimentos = sessao.movimentos if sessao.movimentos else []
        sessoes_com_movimentos.append({
            "sessao": sessao,
            "movimentos": list(movimentos),
            "sessao_id": sessao.id
        })

    paginator = Paginator(sessoes_com_movimentos, 10)
    try:
        sessoes_paginadas = paginator.page(page)
    except:
        sessoes_paginadas = paginator.page(1)

    return render(request, 'contagens/sessoes.html', {
        "sessoes": sessoes_paginadas,
        "current_sort": sort_field,
        "pesquisadores": pesquisadores,
        "codigos": codigos,
        "pontos": pontos,
        "datas": datas,
        "padroes": padroes,
        "filtros_ativos": {
            "status": status,
            "pesquisador": pesquisador,
            "codigo": codigo,
            "ponto": ponto,
            "data": data,
            "padrao": padrao,
            "id": id_sessao
        }
    })


def detalhes_sessao(request, sessao_id):
    try:
        sessao = Session.objects.get(id=sessao_id)
    except Session.DoesNotExist:
        return render(request, 'contagens/detalhes_sessao.html', {'erro': 'Sessão não encontrada.'})
    
    try:
        padrao_types = list(PadraoContagem.objects.values_list('pattern_type', flat=True).distinct())
        
        pattern_type = sessao.padrao if sessao.padrao else (padrao_types[0] if padrao_types else None)
        
        padroes_ordenados = PadraoContagem.objects.filter(
            pattern_type=pattern_type
        ).order_by('order')
        
        if not padroes_ordenados.exists() and padrao_types:
            pattern_type = padrao_types[0]
            padroes_ordenados = PadraoContagem.objects.filter(
                pattern_type=pattern_type
            ).order_by('order')
        
        # Lista de veículos na ordem correta
        veiculos_ordenados = [padrao.veiculo for padrao in padroes_ordenados]
        
        # Obter contagens da sessão
        contagens = Counting.objects.filter(sessao=sessao).order_by("periodo", "movimento", "veiculo")
        
        # Adicionar veículos das contagens que podem não estar nos padrões
        for contagem in contagens:
            if contagem.veiculo not in veiculos_ordenados:
                veiculos_ordenados.append(contagem.veiculo)
        
        # Determinar todos os movimentos (da sessão e das contagens)
        movimentos = []
        if sessao.movimentos:
            movimentos.extend(sessao.movimentos)
        movimentos.extend([c.movimento for c in contagens])
        movimentos = sorted(set(movimentos))

        
        # Determinar todos os períodos
        periodos = set()
        for contagem in contagens:
            if contagem.periodo:
                periodos.add(contagem.periodo)
        
        # Se não houver períodos, criar pelo menos um período baseado no horário inicial
        if not periodos:
            horario_inicio = datetime.strptime(sessao.horario_inicio, "%H:%M") if sessao.horario_inicio else datetime.strptime("00:00", "%H:%M")
            periodos.add(horario_inicio.strftime("%H:%M"))
        
        # Organizar períodos em ordem cronológica
        periodos = sorted(periodos)
        
        # Estrutura final de dados
        dados_sessao = {}
        
        # Para cada movimento, inicializar a estrutura
        for movimento in movimentos:
            dados_sessao[movimento] = []
            
            # Para cada período, criar um registro
            for periodo in periodos:
                periodo_datetime = datetime.strptime(periodo, "%H:%M")
                periodo_fim = (periodo_datetime + timedelta(minutes=15)).strftime("%H:%M")
                
                # Inicializar contagens com zero para todos os veículos
                contagens_movimento = {veiculo: 0 for veiculo in veiculos_ordenados}
                
                # Atualizar apenas os veículos que têm contagens neste período e movimento
                for contagem in contagens:
                    if contagem.periodo == periodo and contagem.movimento == movimento:
                        contagens_movimento[contagem.veiculo] = contagem.contagem
                
                # Adicionar o registro do período para este movimento
                dados_sessao[movimento].append({
                    "das": periodo,
                    "as": periodo_fim,
                    "observacao": "Nenhuma",
                    "contagens": contagens_movimento
                })
        
        return render(request, 'contagens/detalhes_sessao.html', {
            "sessao": sessao,
            "dados_sessao": dados_sessao,
            "veiculos": veiculos_ordenados
        })
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return render(request, 'contagens/detalhes_sessao.html', {
            "sessao": sessao,
            "dados_sessao": {},
            "veiculos": [],
            "erro": f"Erro ao carregar os dados da sessão: {str(e)}"
        })

@csrf_exempt
def get_countings(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            sessao_nome = data.get("sessao")
            username = data.get("usuario")
            contagens = data.get("contagens", [])

            if not sessao_nome or not username:
                return JsonResponse({"erro": "Dados obrigatórios ausentes"}, status=400)

            sessao = Session.objects.select_related('criado_por').filter(sessao=sessao_nome).first()
            if not sessao:
                return JsonResponse({"erro": "Sessão não encontrada"}, status=404)

            # Verificar períodos
            periodos_set = {item['periodo'] for item in contagens if 'periodo' in item and item['periodo']}
            if not periodos_set:
                return JsonResponse({"erro": "Nenhum período válido encontrado"}, status=400)

            # Extrair veículos e movimentos de forma otimizada
            veiculos_set = {item['veiculo'] for item in contagens if 'veiculo' in item and item['veiculo']}
            movimentos_set = {item['movimento'] for item in contagens if 'movimento' in item and item['movimento']}

            # Usar movimentos da sessão se necessário
            if sessao.movimentos and not movimentos_set:
                movimentos_set.update(sessao.movimentos)

            # Otimização: Buscar padrões em uma única query
            pattern_type = sessao.padrao
            padroes_ordenados = PadraoContagem.objects.filter(pattern_type=pattern_type).order_by('order')
            
            # Adicionar veículos dos padrões
            veiculos_set.update(padrao.veiculo for padrao in padroes_ordenados)

            # Otimização: Usar bulk_create para inserções em lote
            counting_objects = []
            for periodo in periodos_set:
                # Remover contagens existentes para este período
                Counting.objects.filter(sessao=sessao, periodo=periodo).delete()
                
                # Preparar objetos para bulk_create
                for veiculo in veiculos_set:
                    for movimento in movimentos_set:
                        # Encontrar contagem correspondente
                        contagem_valor = next(
                            (item.get('count', 0) for item in contagens 
                             if item.get('veiculo') == veiculo 
                             and item.get('movimento') == movimento 
                             and item.get('periodo') == periodo),
                            0
                        )
                        
                        counting_objects.append(Counting(
                            sessao=sessao,
                            veiculo=veiculo,
                            movimento=movimento,
                            contagem=contagem_valor,
                            periodo=periodo
                        ))

            # Inserir todos os objetos de uma vez
            Counting.objects.bulk_create(counting_objects)

            # Remover compressão da resposta
            response_data = {
                "mensagem": f"Contagens processadas com sucesso! {len(counting_objects)} contagens criadas."
            }
            
            return JsonResponse(response_data, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"erro": "JSON inválido"}, status=400)
        except Exception as e:
            return JsonResponse({"erro": f"Erro ao processar contagens: {str(e)}"}, status=500)

    return JsonResponse({"erro": "Método não permitido"}, status=405)

@csrf_exempt
def finalizar_sessao(request):
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Método não permitido'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        logging.info(f"[DEBUG] Dados recebidos para finalização: {json.dumps(data, indent=2)}")
        
        sessao_id = data.get('sessao_id')
        
        if not sessao_id:
            return JsonResponse({
                'status': 'error',
                'message': 'ID da sessão não fornecido'
            }, status=400)
        
        try:
            sessao = Session.objects.get(id=sessao_id)
        except Session.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': f'Sessão com ID {sessao_id} não encontrada'
            }, status=404)
        
        # Para API, não verificamos o usuário logado, pois a autenticação é por token
        # Na interface web, isso é verificado pelo Django automaticamente
        
        if sessao.status == "Concluída":
            return JsonResponse({
                'status': 'success',
                'message': 'Esta sessão já está finalizada'
            }, status=200)
        
        from datetime import datetime
        
        sessao.status = "Concluída"
        # Manter o horario_fim original (do planejamento) - não sobrescrever
        # O horário real de finalização será registrado no updated_at
        sessao.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Sessão finalizada com sucesso'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': 'Dados inválidos'
        }, status=400)
    except Exception as e:
        logging.error(f"[ERRO] Erro ao finalizar sessão: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def exportar_csv(request, sessao_id):
    try:
        sessao = get_object_or_404(Session, id=sessao_id)
        
        # Verificar se o usuário tem permissão para exportar a sessão
        if sessao.criado_por != request.user and not request.user.is_staff:
            return HttpResponse('Você não tem permissão para exportar esta sessão', status=403)
        
        # Obter os tipos de padrões disponíveis no banco de dados
        padrao_types = list(PadraoContagem.objects.values_list('pattern_type', flat=True).distinct())
        
        # Obter o tipo de padrão usado na sessão ou usar o primeiro disponível como fallback
        pattern_type = sessao.padrao if sessao.padrao else (padrao_types[0] if padrao_types else None)
        
        # Obter padrões ordenados com base no tipo de padrão da sessão
        padroes_ordenados = PadraoContagem.objects.filter(
            pattern_type=pattern_type
        ).order_by('order')
        
        # Se não encontrar padrões para o tipo específico, tenta com o primeiro disponível
        if not padroes_ordenados.exists() and padrao_types:
            pattern_type = padrao_types[0]
            padroes_ordenados = PadraoContagem.objects.filter(
                pattern_type=pattern_type
            ).order_by('order')
        
        # Lista de veículos na ordem correta dos padrões
        veiculos_ordenados = [padrao.veiculo for padrao in padroes_ordenados]
        
        # Adicionar veículos das contagens que podem não estar nos padrões
        contagens = Counting.objects.filter(sessao=sessao)
        for contagem in contagens:
            if contagem.veiculo not in veiculos_ordenados:
                veiculos_ordenados.append(contagem.veiculo)
        
        # Criar um novo workbook
        wb = openpyxl.Workbook()
        
        # Configurar estilos
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        # Primeira aba - Detalhes da Sessão
        ws_info = wb.active
        ws_info.title = 'Detalhes da Sessão'
        
        # Adicionar informações da sessão
        info_headers = [
            ['ID da Sessão', sessao.id],
            ['Nome da Sessão', sessao.sessao],
            ['Código', sessao.codigo],
            ['Ponto', sessao.ponto],
            ['Data', sessao.data],
            ['Horário', sessao.horario_inicio],
            ['Status', sessao.status],
            ['Criado Por', sessao.criado_por.username],
            ['Email do Pesquisador', sessao.criado_por.email],
            ['Data de Criação', sessao.created_at.strftime('%d/%m/%Y %H:%M:%S') if sessao.created_at else 'N/A'],
            ['Data de Finalização', sessao.updated_at.strftime('%d/%m/%Y %H:%M:%S') if sessao.status == "Concluída" and sessao.updated_at else 'N/A'],
        ]
        
        for row_idx, (header, value) in enumerate(info_headers, 1):
            ws_info.cell(row=row_idx, column=1, value=header).font = header_font
            ws_info.cell(row=row_idx, column=2, value=value)
        
        # Ajustar largura das colunas
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 30
        
        # Obter contagens agrupadas por movimento
        contagens = Counting.objects.filter(sessao=sessao).order_by('periodo', 'movimento', 'veiculo')
        
        # Obter todos os movimentos da sessão
        movimentos = []
        if sessao.movimentos:
            movimentos.extend(sessao.movimentos)
        # Adicionar movimentos das contagens que podem não estar na sessão
        movimentos.extend([c.movimento for c in contagens])
        # Remover duplicatas e ordenar
        movimentos = sorted(set(movimentos))
        
        # Para cada movimento, criar uma aba separada
        for movimento in movimentos:
            ws_mov = wb.create_sheet(f'Contagens - {movimento}')
            
            # Escrever cabeçalho
            headers = ['Período'] + veiculos_ordenados
            for col_idx, header in enumerate(headers, 1):
                cell = ws_mov.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Agrupar contagens por período
            periodos = defaultdict(lambda: defaultdict(int))
            for contagem in contagens:
                if contagem.movimento == movimento:
                    periodos[contagem.periodo][contagem.veiculo] = contagem.contagem
            
            # Garantir que todos os períodos existam, mesmo sem contagens
            periodos_registrados = set(periodos.keys())
            if not periodos_registrados:
                # Se não houver períodos, crie pelo menos um período baseado no horário inicial
                horario_inicio = datetime.strptime(sessao.horario_inicio, "%H:%M") if sessao.horario_inicio else datetime.strptime("00:00", "%H:%M")
                periodo = horario_inicio.strftime("%H:%M")
                periodos[periodo] = defaultdict(int)
            
            # Escrever dados
            for row_idx, (periodo, contagens_periodo) in enumerate(sorted(periodos.items()), 2):
                ws_mov.cell(row=row_idx, column=1, value=periodo)
                for col_idx, veiculo in enumerate(veiculos_ordenados, 2):
                    ws_mov.cell(row=row_idx, column=col_idx, value=contagens_periodo[veiculo])
            
            # Ajustar largura das colunas
            for col_idx, header in enumerate(headers, 1):
                ws_mov.column_dimensions[get_column_letter(col_idx)].width = max(15, len(header) + 2)

        # Gerar nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sessao_{sessao_id}_{sessao.codigo}_{sessao.ponto}_{timestamp}.xlsx"

        # Tentar salvar no diretório da rede
        try:
            from frontend.utils.config import get_excel_dir
            excel_dir = get_excel_dir()
            
            # Criar subdiretórios por ano e mês
            data_sessao = datetime.strptime(sessao.data, "%Y-%m-%d")
            ano_dir = excel_dir / str(data_sessao.year)
            mes_dir = ano_dir / f"{data_sessao.month:02d}"
            
            # Criar diretórios se não existirem
            mes_dir.mkdir(parents=True, exist_ok=True)
            
            # Caminho completo do arquivo
            file_path = mes_dir / filename
            
            # Salvar o arquivo
            wb.save(file_path)
            
            # Retornar o arquivo como resposta
            response = FileResponse(
                open(file_path, 'rb'),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            print(f"Erro ao salvar no diretório da rede: {str(e)}")
            # Se falhar ao salvar na rede, retorna diretamente como resposta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        
    except Exception as e:
        return HttpResponse(f'Erro ao exportar sessão: {str(e)}', status=500)

@csrf_exempt
def registrar_sessao(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            sessao_nome = data.get("sessao")
            codigo = data.get("codigo")
            ponto = data.get("ponto")
            data_sessao = data.get("data")
            horario_inicio = data.get("horario_inicio")
            horario_fim = data.get("horario_fim")
            username = data.get("usuario")
            status_sessao = data.get("status", "Em andamento")
            movimentos = data.get("movimentos")
            padrao = data.get("padrao", "")

            if not all([sessao_nome, codigo, ponto, data_sessao, horario_inicio, username]):
                return JsonResponse({"erro": "Dados obrigatórios ausentes"}, status=400)

            usuario = User.objects.filter(username=username).first()
            if not usuario:
                return JsonResponse({"erro": "Usuário não encontrado"}, status=404)

            sessao, created = Session.objects.get_or_create(
                sessao=sessao_nome,
                defaults={
                    "codigo": codigo,
                    "ponto": ponto,
                    "data": data_sessao,
                    "horario_inicio": horario_inicio,
                    "horario_fim": horario_fim,
                    "criado_por": usuario,
                    "status": status_sessao,
                    "movimentos": movimentos,
                    "padrao": padrao
                }
            )

            if not created:
                # Atualiza se já existe
                sessao.codigo = codigo
                sessao.ponto = ponto
                sessao.data = data_sessao
                sessao.horario_inicio = horario_inicio
                sessao.status = status_sessao
                sessao.movimentos = movimentos
                sessao.padrao = padrao
                sessao.save()

            return JsonResponse({
                "mensagem": "Sessão registrada com sucesso",
                "id": sessao.id,
                "sessao": sessao.sessao,
                "status": sessao.status
            }, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"erro": "JSON inválido"}, status=400)
        except Exception as e:
            import traceback
            logging.error(f"[ERRO] Exceção em registrar_sessao: {e}")
            logging.error(f"[ERRO] Traceback: {traceback.format_exc()}")
            return JsonResponse({"erro": str(e)}, status=500)
    else:
        return JsonResponse({"erro": "Método não permitido"}, status=405)

def buscar_sessao(request):
    """Busca uma sessão pelo nome e retorna seu ID numérico"""
    if request.method != 'GET':
        return JsonResponse({"erro": "Método não permitido"}, status=405)
    
    try:
        nome_sessao = request.GET.get('nome')
        if not nome_sessao:
            return JsonResponse({"erro": "Nome da sessão não fornecido"}, status=400)
        
        sessao = Session.objects.filter(sessao=nome_sessao).first()
        if not sessao:
            return JsonResponse({"erro": f"Sessão '{nome_sessao}' não encontrada"}, status=404)
        
        return JsonResponse({
            "id": sessao.id,
            "sessao": sessao.sessao,
            "status": sessao.status
        })
    
    except Exception as e:
        logging.error(f"[ERRO] Erro ao buscar sessão por nome: {e}")
        return JsonResponse({"erro": str(e)}, status=500)

@csrf_exempt
def finalizar_por_nome(request):
    """Finaliza uma sessão usando o nome em vez do ID"""
    if request.method != 'POST':
        return JsonResponse({"erro": "Método não permitido"}, status=405)
    
    try:
        data = json.loads(request.body)
        sessao_nome = data.get('sessao_nome')
        
        if not sessao_nome:
            return JsonResponse({"erro": "Nome da sessão não fornecido"}, status=400)
        
        sessao = Session.objects.filter(sessao=sessao_nome).first()
        if not sessao:
            return JsonResponse({"erro": f"Sessão '{sessao_nome}' não encontrada"}, status=404)
        
        if sessao.status == "Concluída":
            return JsonResponse({
                "status": "success",
                "message": "Esta sessão já está finalizada"
            })
        
        from datetime import datetime
        
        sessao.status = "Concluída"
        # Manter o horario_fim original (do planejamento) - não sobrescrever
        # O horário real de finalização será registrado no updated_at
        sessao.save()
        
        return JsonResponse({
            "status": "success",
            "message": f"Sessão '{sessao_nome}' finalizada com sucesso"
        })
    
    except json.JSONDecodeError:
        return JsonResponse({"erro": "JSON inválido"}, status=400)
    except Exception as e:
        logging.error(f"[ERRO] Erro ao finalizar sessão por nome: {e}")
        return JsonResponse({"erro": str(e)}, status=500)



@require_http_methods(["GET"])
def get_pontos_por_codigo(request):
    """
    API endpoint para buscar pontos baseados no código selecionado
    """
    codigo = request.GET.get('codigo', '')
    
    if not codigo:
        return JsonResponse({'pontos': []})
    
    # Buscar pontos únicos para o código específico
    pontos = Session.objects.filter(codigo=codigo).values_list('ponto', flat=True).distinct().order_by('ponto')
    
    return JsonResponse({
        'pontos': list(pontos)
    })

@csrf_exempt
@require_http_methods(["POST"])
def atualizar_contagens(request):
    """
    API endpoint para atualizar contagens de uma sessão
    Usado pela aba de edição de períodos
    """
    print("[DJANGO DEBUG] === FUNÇÃO atualizar_contagens CHAMADA ===")
    try:
        data = json.loads(request.body)
        sessao_nome = data.get('sessao')
        usuario = data.get('usuario')
        contagens = data.get('contagens', [])
        
        print(f"[DJANGO DEBUG] Dados recebidos: sessao={sessao_nome}, usuario={usuario}, contagens={len(contagens)}")
        logging.info(f"[DEBUG] Recebido request para atualizar contagens: sessao={sessao_nome}, usuario={usuario}, contagens={len(contagens)}")
        
        if not sessao_nome:
            print("[DJANGO ERROR] Nome da sessão é obrigatório")
            logging.error("[ERROR] Nome da sessão é obrigatório")
            return JsonResponse({'erro': 'Nome da sessão é obrigatório'}, status=400)
        
        # Buscar sessão pelo nome
        try:
            sessao = Session.objects.get(sessao=sessao_nome)
            print(f"[DJANGO DEBUG] Sessão encontrada: ID={sessao.id}, Nome={sessao.sessao}")
            logging.info(f"[DEBUG] Sessão encontrada: ID={sessao.id}, Nome={sessao.sessao}")
        except Session.DoesNotExist:
            print(f"[DJANGO ERROR] Sessão {sessao_nome} não encontrada")
            logging.error(f"[ERROR] Sessão {sessao_nome} não encontrada")
            return JsonResponse({'erro': f'Sessão {sessao_nome} não encontrada'}, status=404)
        
        # Log da operação
        print(f"[DJANGO INFO] Atualizando contagens para sessão {sessao_nome} por {usuario}")
        logging.info(f"[INFO] Atualizando contagens para sessão {sessao_nome} por {usuario}")
        
        # Verificar quantos registros existem atualmente para esta sessão
        existing_count = Counting.objects.filter(sessao=sessao).count()
        print(f"[DJANGO DEBUG] Registros existentes para sessão {sessao_nome}: {existing_count}")
        
        # Se há registros existentes, mostrar alguns exemplos
        if existing_count > 0:
            sample_records = Counting.objects.filter(sessao=sessao)[:5]
            for record in sample_records:
                print(f"[DJANGO DEBUG] Registro existente: {record.veiculo}-{record.movimento}-{record.periodo}: {record.contagem}")
        
        # Atualizar ou criar contagens
        contagens_atualizadas = 0
        contagens_criadas = 0
        duplicatas_removidas = 0
        
        print(f"[DJANGO DEBUG] Processando {len(contagens)} contagens...")
        
        for contagem_data in contagens:
            veiculo = contagem_data.get('veiculo')
            movimento = contagem_data.get('movimento')
            count = contagem_data.get('count', 0)
            periodo = contagem_data.get('periodo')
            
            print(f"[DJANGO DEBUG] Processando: veiculo={veiculo}, movimento={movimento}, count={count}, periodo={periodo}")
            logging.debug(f"[DEBUG] Processando: veiculo={veiculo}, movimento={movimento}, count={count}, periodo={periodo}")
            
            if not all([veiculo, movimento, periodo is not None]):
                logging.warning(f"[WARNING] Dados incompletos: veiculo={veiculo}, movimento={movimento}, periodo={periodo}")
                continue
            
            # Buscar ou criar contagem
            try:
                # Primeiro, encontrar TODOS os registros existentes (pode haver duplicatas)
                existing_records = Counting.objects.filter(
                    sessao=sessao,
                    veiculo=veiculo,
                    movimento=movimento,
                    periodo=periodo
                )
                
                record_count = existing_records.count()
                print(f"[DJANGO DEBUG] Encontrados {record_count} registros para {veiculo}-{movimento}-{periodo}")
                
                if record_count == 0:
                    # Nenhum registro existe - criar novo
                    contagem_obj = Counting.objects.create(
                        sessao=sessao,
                        veiculo=veiculo,
                        movimento=movimento,
                        periodo=periodo,
                        contagem=count
                    )
                    contagens_criadas += 1
                    print(f"[DJANGO DEBUG] Contagem CRIADA: {veiculo}-{movimento}-{periodo}: {count}")
                    
                elif record_count == 1:
                    # Exatamente um registro - atualizar normalmente
                    contagem_obj = existing_records.first()
                    if contagem_obj.contagem != count:
                        old_count = contagem_obj.contagem
                        contagem_obj.contagem = count
                        contagem_obj.save()
                        contagens_atualizadas += 1
                        print(f"[DJANGO DEBUG] Contagem ATUALIZADA: {veiculo}-{movimento}-{periodo}: {old_count} -> {count}")
                    else:
                        print(f"[DJANGO DEBUG] Contagem INALTERADA: {veiculo}-{movimento}-{periodo}: {count}")
                        
                else:
                    # MÚLTIPLOS registros (duplicatas) - limpar e manter apenas um
                    print(f"[DJANGO WARNING] Encontradas {record_count} duplicatas para {veiculo}-{movimento}-{periodo}")
                    
                    # Manter o mais recente (com maior ID)
                    latest_record = existing_records.order_by('-id').first()
                    
                    # Deletar os outros
                    duplicates_to_delete = existing_records.exclude(id=latest_record.id)
                    deleted_count = duplicates_to_delete.count()
                    duplicates_to_delete.delete()
                    duplicatas_removidas += deleted_count
                    
                    print(f"[DJANGO DEBUG] Removidas {deleted_count} duplicatas, mantido registro ID={latest_record.id}")
                    
                    # Agora atualizar o registro mantido
                    if latest_record.contagem != count:
                        old_count = latest_record.contagem
                        latest_record.contagem = count
                        latest_record.save()
                        contagens_atualizadas += 1
                        print(f"[DJANGO DEBUG] Contagem ATUALIZADA (após limpeza): {veiculo}-{movimento}-{periodo}: {old_count} -> {count}")
                    else:
                        print(f"[DJANGO DEBUG] Contagem INALTERADA (após limpeza): {veiculo}-{movimento}-{periodo}: {count}")
                    
            except Exception as db_ex:
                print(f"[DJANGO ERROR] Erro ao processar contagem {veiculo}-{movimento}-{periodo}: {str(db_ex)}")
                logging.error(f"[ERROR] Erro ao processar contagem {veiculo}-{movimento}-{periodo}: {str(db_ex)}")
                continue
        
        resultado = {
            'message': 'Contagens atualizadas com sucesso',
            'detalhes': {
                'contagens_criadas': contagens_criadas,
                'contagens_atualizadas': contagens_atualizadas,
                'duplicatas_removidas': duplicatas_removidas,
                'total_processadas': len(contagens)
            }
        }
        
        print(f"[DJANGO SUCCESS] Resultado: {resultado}")
        logging.info(f"[SUCCESS] Resultado: {resultado}")
        return JsonResponse(resultado)
        
    except json.JSONDecodeError as json_ex:
        print(f"[DJANGO ERROR] JSON inválido: {str(json_ex)}")
        logging.error(f"[ERROR] JSON inválido: {str(json_ex)}")
        return JsonResponse({'erro': 'JSON inválido'}, status=400)
    except Exception as e:
        print(f"[DJANGO ERROR] Erro ao atualizar contagens: {str(e)}")
        logging.error(f"[ERROR] Erro ao atualizar contagens: {str(e)}")
        import traceback
        traceback_str = traceback.format_exc()
        print(f"[DJANGO ERROR] Traceback: {traceback_str}")
        logging.error(f"[ERROR] Traceback: {traceback_str}")
        return JsonResponse({'erro': f'Erro interno: {str(e)}'}, status=500)
